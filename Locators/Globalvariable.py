class Globalvariable():
    import openpyxl
    path = "C:\\Users\\Anupam\\Desktop\\Testing.xlsx"
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    baseURL ="https://www.phptravels.net/login"
    username = 'ragnar@gmail.com'
    password = "zxcv@1234"

    #Signup page

    firstname = sheet.cell(2, 1).value
    lastname = sheet.cell(2, 2).value
    mobilenum = sheet.cell(2, 3).value
    emailid = sheet.cell(2, 4).value
    pwd = sheet.cell(2, 5).value
    confirmpwd = sheet.cell(2, 6).value